import openpyxl

#https://www.geeksforgeeks.org/read-json-file-using-python/

book = openpyxl.load_workbook("F:\\PythonSelenium\\dataSheet.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
dict ={}

sheet.cell(row=2, column=2).value = "nimmaarun@gmail.com"

print(sheet.cell(row=2, column=2).value)
print(sheet['B3'].value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "testcase2":
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            #print(sheet.cell(row=i, column=j).value)

print(dict)