import openpyxl


class HomePageDataClass:

    test_Home_page_Data = [{"email":"nimmaarun@gmail.com", "pwd":"Password" , "gender":"Female"} ,
                            {"email":"ni@gmail.com", "pwd":"Pas" , "gender":"Male"}]

    @staticmethod
    def getTestData(testcasename):
        book = openpyxl.load_workbook("F:\\PythonSelenium\\dataSheet.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcasename:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    # print(sheet.cell(row=i, column=j).value)

        return[dict]